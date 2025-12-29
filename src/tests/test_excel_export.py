"""Excel export tests."""

from pathlib import Path
import tempfile

from openpyxl import load_workbook

from mcp_analyst.exports.excel_dcf import export_dcf_to_excel
from mcp_analyst.orchestrator.run_context import RunContext
from mcp_analyst.schemas.valuation import (
    DcfAssumptions,
    DcfResults,
    ValuationOutput,
)


def test_excel_export_creates_file():
    """Test that Excel export creates a file."""
    with tempfile.TemporaryDirectory() as tmpdir:
        run_context = RunContext.create(
            ticker="TEST",
            sector=None,
            horizon="5y",
            risk="moderate",
            focus=None,
            terminal="gordon",
            output_dir=Path(tmpdir),
        )

        assumptions = DcfAssumptions(
            horizon_years=5,
            terminal_method="gordon",
            wacc=0.10,
            terminal_growth_rate=0.03,
            revenue_growth_rates=[0.15, 0.12],
            margin_assumptions={"operating_margin": 0.15},
        )
        results = DcfResults(
            fair_value_per_share=50.0,
            total_enterprise_value=1000000000.0,
            equity_value=1000000000.0,
            present_values={},
        )
        valuation_output = ValuationOutput(
            assumptions=assumptions, results=results
        )

        excel_path = export_dcf_to_excel(valuation_output, run_context)

        assert excel_path.exists()
        assert excel_path.suffix == ".xlsx"


def test_excel_export_has_required_sheets():
    """Test that Excel export has required sheets."""
    with tempfile.TemporaryDirectory() as tmpdir:
        run_context = RunContext.create(
            ticker="TEST",
            sector=None,
            horizon="5y",
            risk="moderate",
            focus=None,
            terminal="gordon",
            output_dir=Path(tmpdir),
        )

        assumptions = DcfAssumptions(
            horizon_years=5,
            terminal_method="gordon",
            wacc=0.10,
            terminal_growth_rate=0.03,
            revenue_growth_rates=[],
            margin_assumptions={},
        )
        results = DcfResults(
            fair_value_per_share=50.0,
            total_enterprise_value=1000000000.0,
            equity_value=1000000000.0,
            present_values={},
        )
        valuation_output = ValuationOutput(
            assumptions=assumptions, results=results
        )

        excel_path = export_dcf_to_excel(valuation_output, run_context)

        wb = load_workbook(excel_path)
        sheet_names = wb.sheetnames
        assert "Assumptions" in sheet_names
        assert "Results" in sheet_names

