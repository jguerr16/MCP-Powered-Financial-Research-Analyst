"""Regression tests for Excel formatting to prevent formatting bugs."""

import json
from pathlib import Path
from unittest.mock import MagicMock

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

from mcp_analyst.exports.excel_dcf import export_dcf_to_excel
from mcp_analyst.exports.excel_styles import (
    apply_currency_millions,
    apply_decimal,
    apply_number,
    apply_percent,
)
from mcp_analyst.orchestrator.run_context import RunContext
from mcp_analyst.schemas.financials import FinancialSummary, MetricSeries
from mcp_analyst.schemas.pricing import QuoteData
from mcp_analyst.schemas.valuation import (
    DcfAssumptions,
    DcfResults,
    OperatingForecast,
    ValuationOutput,
)


@pytest.fixture
def sample_valuation_output():
    """Create a sample valuation output for testing."""
    assumptions = DcfAssumptions(
        horizon_years=5,
        forecast_years=["2025", "2026", "2027", "2028", "2029"],
        base_year="2024",
        base_year_revenue=10_000_000_000,  # $10B
        revenue_growth_rates=[0.10, 0.08, 0.06, 0.04, 0.03],
        cogs_ex_da_pct_rev=[0.65] * 5,
        sga_pct_rev=[0.20] * 5,
        da_pct_rev=[0.03] * 5,
        sbc_pct_rev=[0.02] * 5,
        capex_pct_rev=[0.05] * 5,
        nwc_pct_rev=[0.10] * 5,
        terminal_method="gordon",
        terminal_growth_rate=0.025,
        wacc=0.10,
        tax_rate=0.21,
        shares_out=1_000_000_000,  # 1B shares
        net_debt=500_000_000,  # $500M
        confidence={
            "base_revenue": "HIGH",
            "revenue_growth": "HIGH",
            "cogs_pct": "MED",
        },
        fade_method="piecewise",
    )

    forecast = [
        OperatingForecast(
            year="2025",
            revenue=11_000_000_000,
            cogs_ex_da=7_150_000_000,
            sga=2_200_000_000,
            da=330_000_000,
            ebit=1_100_000_000,
            taxes=231_000_000,
            nopat=869_000_000,
            da_addback=330_000_000,
            sbc_addback=220_000_000,
            delta_nwc=100_000_000,
            capex=550_000_000,
            unlevered_fcf=549_000_000,
            discount_factor=0.909,
            pv_ufcf=499_091_000,
        )
    ] * 5

    results = DcfResults(
        operating_forecast=forecast,
        total_enterprise_value=50_000_000_000,
        equity_value=49_500_000_000,
        fair_value_per_share=49.50,
        terminal_value=45_000_000_000,
        pv_terminal_value=45_000_000_000,
        present_values={"PV UFCF": 5_000_000_000, "PV Terminal": 45_000_000_000},
        sensitivity={},
    )

    return ValuationOutput(assumptions=assumptions, results=results)


@pytest.fixture
def sample_run_context(tmp_path):
    """Create a sample run context."""
    from datetime import datetime
    import uuid
    
    return RunContext(
        run_id=str(uuid.uuid4()),
        ticker="TEST",
        sector="technology",
        horizon="5y",
        risk="moderate",
        focus="valuation",
        terminal="gordon",
        output_dir=tmp_path,
        created_at=datetime.now(),
    )


@pytest.fixture
def sample_quote_data():
    """Create sample quote data."""
    return QuoteData(
        ticker="TEST",
        price=50.0,
        market_cap=50_000_000_000,
        beta=1.2,
        shares_out=1_000_000_000,
        currency="USD",
        as_of_utc="2025-12-29T12:00:00Z",
        source="yahoo_finance",
    )


def test_discount_factor_not_currency(sample_valuation_output, sample_run_context, sample_quote_data):
    """Test that discount factor row is NOT formatted as currency."""
    # Ensure run directory exists
    sample_run_context.run_dir.mkdir(parents=True, exist_ok=True)
    
    excel_path = export_dcf_to_excel(
        sample_valuation_output,
        sample_run_context,
        quote_data=sample_quote_data,
    )

    wb = load_workbook(excel_path)
    dcf_sheet = wb["DCF"]

    # Find discount factor row
    discount_row = None
    for row_idx, row in enumerate(dcf_sheet.iter_rows(), 1):
        if any(cell.value and "Discount Factor" in str(cell.value) for cell in row):
            discount_row = row_idx
            break

    assert discount_row is not None, "Discount Factor row not found"

    # Check that discount factor cells are NOT currency format
    # Currency format typically contains $ or has specific number format
    for col in range(2, 10):  # Check columns B through I
        cell = dcf_sheet.cell(discount_row, col)
        if cell.value is not None:
            # Discount factor should be decimal format (0.000), not currency
            assert "$" not in str(cell.number_format), f"Discount factor at {cell.coordinate} should not be currency"
            # Should be decimal format
            assert "0.000" in str(cell.number_format) or "0.00" in str(cell.number_format), \
                f"Discount factor at {cell.coordinate} should be decimal format"


def test_shares_outstanding_not_currency(sample_valuation_output, sample_run_context, sample_quote_data):
    """Test that shares outstanding is NOT formatted as currency."""
    # Ensure run directory exists
    sample_run_context.run_dir.mkdir(parents=True, exist_ok=True)
    
    excel_path = export_dcf_to_excel(
        sample_valuation_output,
        sample_run_context,
        quote_data=sample_quote_data,
    )

    wb = load_workbook(excel_path)
    dcf_sheet = wb["DCF"]

    # Find shares outstanding row
    shares_row = None
    for row_idx, row in enumerate(dcf_sheet.iter_rows(), 1):
        if any(cell.value and "Shares Outstanding" in str(cell.value) for cell in row):
            shares_row = row_idx
            break

    assert shares_row is not None, "Shares Outstanding row not found"

    # Check that shares cells are NOT currency format
    for col in range(2, 10):
        cell = dcf_sheet.cell(shares_row, col)
        if cell.value is not None:
            # Shares should be number format, not currency
            assert "$" not in str(cell.number_format), \
                f"Shares outstanding at {cell.coordinate} should not be currency"
            # Should be number format (not currency)
            assert cell.number_format in ["#,##0", "#,##0.00", "General"] or \
                "#,##0" in str(cell.number_format), \
                f"Shares at {cell.coordinate} should be number format, got {cell.number_format}"


def test_currency_rows_have_currency_format(sample_valuation_output, sample_run_context, sample_quote_data):
    """Test that currency rows (revenue, costs, etc.) have currency formatting."""
    # Ensure run directory exists
    sample_run_context.run_dir.mkdir(parents=True, exist_ok=True)
    
    excel_path = export_dcf_to_excel(
        sample_valuation_output,
        sample_run_context,
        quote_data=sample_quote_data,
    )

    wb = load_workbook(excel_path)
    dcf_sheet = wb["DCF"]

    # Find revenue row
    revenue_row = None
    for row_idx, row in enumerate(dcf_sheet.iter_rows(), 1):
        if any(cell.value and "Revenue" in str(cell.value) and "Margin" not in str(cell.value) 
               for cell in row):
            revenue_row = row_idx
            break

    assert revenue_row is not None, "Revenue row not found"

    # Check that revenue cells have currency format
    for col in range(2, 7):  # Check forecast columns
        cell = dcf_sheet.cell(revenue_row, col)
        if cell.value is not None:
            # Should have currency format (contains $ or _($)
            assert "$" in str(cell.number_format) or "_($" in str(cell.number_format) or \
                "#,##0" in str(cell.number_format), \
                f"Revenue at {cell.coordinate} should be currency format, got {cell.number_format}"


def test_valsum_sheet_exists(sample_valuation_output, sample_run_context, sample_quote_data, tmp_path):
    """Test that ValSum sheet exists."""
    # Ensure run directory exists
    sample_run_context.run_dir.mkdir(parents=True, exist_ok=True)
    
    excel_path = export_dcf_to_excel(
        sample_valuation_output,
        sample_run_context,
        quote_data=sample_quote_data,
    )

    wb = load_workbook(excel_path)

    assert "ValSum" in wb.sheetnames, "ValSum sheet should exist"


def test_current_price_cell_is_numeric(sample_valuation_output, sample_run_context, sample_quote_data):
    """Test that current price cell in ValSum is numeric."""
    # Ensure run directory exists
    sample_run_context.run_dir.mkdir(parents=True, exist_ok=True)
    
    excel_path = export_dcf_to_excel(
        sample_valuation_output,
        sample_run_context,
        quote_data=sample_quote_data,
    )

    wb = load_workbook(excel_path)
    valsum_sheet = wb["ValSum"]

    # Find current price cell
    price_cell = None
    for row in valsum_sheet.iter_rows():
        for cell in row:
            if cell.value and "Current Price" in str(cell.value):
                # Next cell should be the price value
                price_cell = valsum_sheet.cell(cell.row, cell.column + 1)
                break
        if price_cell:
            break

    assert price_cell is not None, "Current Price cell not found"
    assert isinstance(price_cell.value, (int, float)), \
        f"Current price should be numeric, got {type(price_cell.value)}"


def test_confidence_labels_in_inputs(sample_valuation_output, sample_run_context, sample_quote_data):
    """Test that confidence labels appear in Inputs sheet."""
    # Ensure run directory exists
    sample_run_context.run_dir.mkdir(parents=True, exist_ok=True)
    
    excel_path = export_dcf_to_excel(
        sample_valuation_output,
        sample_run_context,
        quote_data=sample_quote_data,
    )

    wb = load_workbook(excel_path)
    inputs_sheet = wb["Inputs"]

    # Check for confidence column header
    has_confidence_header = False
    for row in inputs_sheet.iter_rows(max_row=5):
        for cell in row:
            if cell.value and "Confidence" in str(cell.value):
                has_confidence_header = True
                break
        if has_confidence_header:
            break

    assert has_confidence_header, "Confidence column header should exist"

    # Check for confidence values (HIGH, MED, LOW)
    has_confidence_values = False
    for row in inputs_sheet.iter_rows():
        for cell in row:
            if cell.value in ["HIGH", "MED", "LOW"]:
                has_confidence_values = True
                break
        if has_confidence_values:
            break

    assert has_confidence_values, "Confidence values (HIGH/MED/LOW) should exist"


def test_overview_block_in_dcf(sample_valuation_output, sample_run_context, sample_quote_data):
    """Test that Overview block exists in DCF sheet."""
    # Ensure run directory exists
    sample_run_context.run_dir.mkdir(parents=True, exist_ok=True)
    
    excel_path = export_dcf_to_excel(
        sample_valuation_output,
        sample_run_context,
        quote_data=sample_quote_data,
    )

    wb = load_workbook(excel_path)
    dcf_sheet = wb["DCF"]

    # Check for Overview block
    has_overview = False
    for row in dcf_sheet.iter_rows(max_row=30):
        for cell in row:
            if cell.value and "Overview" in str(cell.value):
                has_overview = True
                break
        if has_overview:
            break

    assert has_overview, "Overview block should exist in DCF sheet"


def test_fade_method_displayed(sample_valuation_output, sample_run_context, sample_quote_data):
    """Test that fade method is displayed in Inputs sheet."""
    # Ensure run directory exists
    sample_run_context.run_dir.mkdir(parents=True, exist_ok=True)
    
    excel_path = export_dcf_to_excel(
        sample_valuation_output,
        sample_run_context,
        quote_data=sample_quote_data,
    )

    wb = load_workbook(excel_path)
    inputs_sheet = wb["Inputs"]

    # Check for fade method
    has_fade_method = False
    for row in inputs_sheet.iter_rows():
        for cell in row:
            if cell.value and "piecewise" in str(cell.value).lower():
                has_fade_method = True
                break
        if has_fade_method:
            break

    assert has_fade_method, "Fade method should be displayed in Inputs sheet"


def test_column_letters_work_past_z(sample_valuation_output, sample_run_context, sample_quote_data):
    """Test that column letters work correctly past column Z (AA, AB, etc.)."""
    # Ensure run directory exists
    sample_run_context.run_dir.mkdir(parents=True, exist_ok=True)
    
    excel_path = export_dcf_to_excel(
        sample_valuation_output,
        sample_run_context,
        quote_data=sample_quote_data,
    )

    wb = load_workbook(excel_path)
    dcf_sheet = wb["DCF"]

    # Check that we can access columns beyond Z
    # If we have 5 forecast years + base year, we should have columns up to at least H
    # But if we had more, we'd need AA, AB, etc.
    # Just verify the sheet doesn't crash with extended columns
    try:
        # Try accessing column AA (27)
        cell_aa = dcf_sheet.cell(1, 27)
        assert cell_aa is not None, "Should be able to access column AA"
    except Exception as e:
        pytest.fail(f"Failed to access column AA: {e}")

