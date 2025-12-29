"""Excel DCF workbook export using openpyxl - Analyst-style DCF model."""

from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font

from mcp_analyst.exports.excel_styles import (
    apply_currency_millions,
    apply_header,
    apply_input,
    apply_percent,
    freeze_panes,
    set_column_widths,
    to_millions,
)
from mcp_analyst.orchestrator.run_context import RunContext
from mcp_analyst.schemas.financials import FinancialSummary
from mcp_analyst.schemas.valuation import ValuationOutput
from mcp_analyst.tools.pricing import fetch_current_price


def export_dcf_to_excel(
    valuation_output: ValuationOutput,
    run_context: RunContext,
    financial_summary: Optional[FinancialSummary] = None,
) -> Path:
    """
    Export DCF assumptions and results to Excel workbook (analyst-style).

    Args:
        valuation_output: Valuation output with assumptions and results
        run_context: Run context for file naming

    Returns:
        Path to created Excel file
    """
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Sheet order: Historical, Inputs, DCF, Cases, WACC, Sensitivities, Summary
    if financial_summary:
        historical_sheet = wb.create_sheet("Historical", 0)
        _write_historical(historical_sheet, financial_summary, run_context.ticker)

    # Inputs/Assumptions sheet
    inputs_sheet = wb.create_sheet("Inputs", 1)
    _write_inputs(inputs_sheet, valuation_output.assumptions, run_context.ticker)

    # DCF Forecast sheet (main model)
    dcf_sheet = wb.create_sheet("DCF", 2)
    _write_dcf_forecast(dcf_sheet, valuation_output, run_context.ticker)

    # Cases sheet
    cases_sheet = wb.create_sheet("Cases", 3)
    _write_cases(cases_sheet, valuation_output.assumptions)

    # WACC Calculation sheet
    wacc_sheet = wb.create_sheet("WACC", 4)
    _write_wacc(wacc_sheet, valuation_output.assumptions)

    # Sensitivities sheet
    sensitivities_sheet = wb.create_sheet("Sensitivities", 5)
    _write_sensitivities(sensitivities_sheet, valuation_output.results.sensitivity)

    # Summary sheet
    summary_sheet = wb.create_sheet("Summary", 6)
    _write_summary(summary_sheet, valuation_output, run_context.ticker)

    # Save workbook
    date_str = run_context.created_at.strftime("%Y-%m-%d")
    filename = f"DCF_{run_context.ticker}_{date_str}.xlsx"
    excel_path = run_context.run_dir / filename
    wb.save(excel_path)

    return excel_path


def _add_title_block(sheet, ticker: str, title: str, row: int = 1) -> int:
    """Add title block to sheet. Returns next row number."""
    sheet.cell(row, 1).value = f"{ticker} - {title}"
    sheet.cell(row, 1).font = Font(bold=True, size=14)
    row += 1
    sheet.cell(row, 1).value = "$ Millions except per share"
    sheet.cell(row, 1).font = Font(italic=True, size=10)
    return row + 2  # Return row after title block


def _write_historical(sheet, financial_summary: FinancialSummary, ticker: str) -> None:
    """Write historical financial data to sheet in $ Millions."""
    row = _add_title_block(sheet, ticker, "Historical Financial Data")

    # Headers
    headers = ["Period"]
    metric_names = [m.metric_name for m in financial_summary.metrics]
    headers.extend(metric_names)

    header_row = row
    for col, header in enumerate(headers, start=1):
        cell = sheet.cell(header_row, col)
        cell.value = header
    apply_header(sheet, f"A{header_row}:{chr(64 + len(headers))}{header_row}")

    # Write data in $ Millions
    periods = financial_summary.periods[:10]  # Last 10 periods
    data_start_row = header_row + 1
    for row_idx, period in enumerate(periods, start=data_start_row):
        sheet.cell(row_idx, 1).value = period

        for col_idx, metric in enumerate(financial_summary.metrics, start=2):
            period_idx = financial_summary.periods.index(period)
            if period_idx < len(metric.values):
                value = metric.values[period_idx]
                # Convert to millions for USD metrics
                if metric.unit == "USD":
                    value = to_millions(value)
                sheet.cell(row_idx, col_idx).value = value

    # Apply formatting
    if financial_summary.metrics:
        num_cols = len(headers)
        apply_currency_millions(sheet, f"B{data_start_row}:{chr(64 + num_cols)}{data_start_row + len(periods) - 1}")

    # Set column widths
    set_column_widths(sheet, {1: 15, **{i: 14 for i in range(2, num_cols + 1)}})

    # Freeze panes
    freeze_panes(sheet, "B3")


def _write_inputs(sheet, assumptions, ticker: str) -> None:
    """Write input assumptions to sheet."""
    row = _add_title_block(sheet, ticker, "Model Inputs")

    # Headers
    sheet.cell(row, 1).value = "Input"
    sheet.cell(row, 2).value = "Value"
    apply_header(sheet, f"A{row}:B{row}")

    row += 1
    input_start_row = row

    inputs = [
        ("Base Year", assumptions.base_year, None),
        ("Base Revenue ($M)", to_millions(assumptions.base_year_revenue), "currency"),
        ("Horizon (years)", assumptions.horizon_years, None),
        ("Tax Rate", assumptions.tax_rate, "percent"),
        ("WACC", assumptions.wacc, "percent"),
        ("Terminal Growth Rate", assumptions.terminal_growth_rate, "percent"),
        ("Terminal Method", assumptions.terminal_method, None),
        ("Shares Outstanding (M)", to_millions(assumptions.shares_out), None),
        ("Net Debt ($M)", to_millions(assumptions.net_debt), "currency"),
    ]

    for label, value, fmt_type in inputs:
        sheet.cell(row, 1).value = label
        sheet.cell(row, 2).value = value
        if fmt_type == "currency":
            apply_currency_millions(sheet, f"B{row}")
        elif fmt_type == "percent":
            apply_percent(sheet, f"B{row}")
        row += 1

    # Revenue growth rates
    sheet.cell(row, 1).value = "Revenue Growth Rates"
    sheet.cell(row, 1).font = Font(bold=True)
    row += 1
    for i, rate in enumerate(assumptions.revenue_growth_rates):
        sheet.cell(row, 1).value = f"Year {i+1}"
        sheet.cell(row, 2).value = rate
        apply_percent(sheet, f"B{row}")
        row += 1

    # Apply input styling
    apply_input(sheet, f"A{input_start_row}:B{row-1}")

    # Set column widths
    set_column_widths(sheet, {1: 30, 2: 15})


def _write_dcf_forecast(sheet, valuation_output: ValuationOutput, ticker: str) -> None:
    """Write full DCF operating forecast to sheet in $ Millions."""
    row = _add_title_block(sheet, ticker, "Discounted Cash Flow Model")

    assumptions = valuation_output.assumptions
    forecast = valuation_output.results.operating_forecast

    # Headers
    headers = ["", "Base Year"] + assumptions.forecast_years + ["Terminal"]
    header_row = row

    for col, header in enumerate(headers, start=1):
        cell = sheet.cell(header_row, col)
        cell.value = header
    apply_header(sheet, f"A{header_row}:{chr(64 + len(headers))}{header_row}")

    row_labels = [
        "Revenue",
        "(-) COGS ex D&A",
        "(-) SG&A",
        "(-) D&A",
        "EBIT",
        "(-) Taxes",
        "NOPAT",
        "(+) D&A add-back",
        "(+) SBC add-back",
        "(-) ΔNWC",
        "(-) Capex",
        "Unlevered FCF",
        "Discount Factor",
        "PV of UFCF",
    ]

    # Write row labels
    data_start_row = header_row + 1
    for row_idx, label in enumerate(row_labels, start=data_start_row):
        sheet.cell(row_idx, 1).value = label

    # Write base year
    sheet.cell(data_start_row, 2).value = to_millions(assumptions.base_year_revenue)

    # Fill in forecast years in $ Millions
    for col_idx, year in enumerate(assumptions.forecast_years, start=3):
        if col_idx - 3 < len(forecast):
            f = forecast[col_idx - 3]
            sheet.cell(data_start_row, col_idx).value = to_millions(f.revenue)  # Revenue
            sheet.cell(data_start_row + 1, col_idx).value = -to_millions(f.cogs_ex_da)  # COGS
            sheet.cell(data_start_row + 2, col_idx).value = -to_millions(f.sga)  # SG&A
            sheet.cell(data_start_row + 3, col_idx).value = -to_millions(f.da)  # D&A
            sheet.cell(data_start_row + 4, col_idx).value = to_millions(f.ebit)  # EBIT
            sheet.cell(data_start_row + 5, col_idx).value = -to_millions(f.taxes)  # Taxes
            sheet.cell(data_start_row + 6, col_idx).value = to_millions(f.nopat)  # NOPAT
            sheet.cell(data_start_row + 7, col_idx).value = to_millions(f.da_addback)  # D&A add-back
            sheet.cell(data_start_row + 8, col_idx).value = to_millions(f.sbc_addback)  # SBC add-back
            sheet.cell(data_start_row + 9, col_idx).value = -to_millions(f.delta_nwc)  # ΔNWC
            sheet.cell(data_start_row + 10, col_idx).value = -to_millions(f.capex)  # Capex
            sheet.cell(data_start_row + 11, col_idx).value = to_millions(f.unlevered_fcf)  # UFCF
            sheet.cell(data_start_row + 12, col_idx).value = f.discount_factor  # Discount factor
            sheet.cell(data_start_row + 13, col_idx).value = to_millions(f.pv_ufcf)  # PV of UFCF

    # Terminal value
    terminal_col = len(assumptions.forecast_years) + 3
    sheet.cell(data_start_row + 13, terminal_col).value = to_millions(valuation_output.results.pv_terminal_value)

    # Apply currency formatting
    num_cols = len(headers)
    apply_currency_millions(sheet, f"B{data_start_row}:{chr(64 + num_cols)}{data_start_row + 13}")

    # Summary rows
    summary_row = data_start_row + len(row_labels) + 2
    sheet.cell(summary_row, 1).value = "Total PV of UFCF"
    sheet.cell(summary_row, 1).font = Font(bold=True)
    sheet.cell(summary_row, 2).value = to_millions(sum(f.pv_ufcf for f in forecast))
    apply_currency_millions(sheet, f"B{summary_row}")

    sheet.cell(summary_row + 1, 1).value = "PV of Terminal Value"
    sheet.cell(summary_row + 1, 1).font = Font(bold=True)
    sheet.cell(summary_row + 1, 2).value = to_millions(valuation_output.results.pv_terminal_value)
    apply_currency_millions(sheet, f"B{summary_row + 1}")

    sheet.cell(summary_row + 2, 1).value = "Enterprise Value"
    sheet.cell(summary_row + 2, 1).font = Font(bold=True, size=12)
    sheet.cell(summary_row + 2, 2).value = to_millions(valuation_output.results.total_enterprise_value)
    sheet.cell(summary_row + 2, 2).font = Font(bold=True, size=12)
    apply_currency_millions(sheet, f"B{summary_row + 2}")

    sheet.cell(summary_row + 3, 1).value = "(-) Net Debt"
    sheet.cell(summary_row + 3, 2).value = -to_millions(assumptions.net_debt)
    apply_currency_millions(sheet, f"B{summary_row + 3}")

    sheet.cell(summary_row + 4, 1).value = "Equity Value"
    sheet.cell(summary_row + 4, 1).font = Font(bold=True)
    sheet.cell(summary_row + 4, 2).value = to_millions(valuation_output.results.equity_value)
    apply_currency_millions(sheet, f"B{summary_row + 4}")

    sheet.cell(summary_row + 5, 1).value = "Shares Outstanding (M)"
    sheet.cell(summary_row + 5, 2).value = to_millions(assumptions.shares_out)

    sheet.cell(summary_row + 6, 1).value = "Fair Value per Share"
    sheet.cell(summary_row + 6, 1).font = Font(bold=True, size=14)
    sheet.cell(summary_row + 6, 2).value = valuation_output.results.fair_value_per_share
    sheet.cell(summary_row + 6, 2).font = Font(bold=True, size=14)
    sheet.cell(summary_row + 6, 2).number_format = '#,##0.00'

    # Set column widths
    set_column_widths(sheet, {1: 30, **{i: 14 for i in range(2, num_cols + 1)}})

    # Freeze panes
    freeze_panes(sheet, "B3")


def _write_cases(sheet, assumptions) -> None:
    """Write scenario cases (Base/Bull/Bear) to sheet."""
    row = 1
    headers = ["Case", "Growth Fade", "Steady Margin", "WACC", "Terminal Growth"]
    for col, header in enumerate(headers, start=1):
        cell = sheet.cell(row, col)
        cell.value = header
    apply_header(sheet, f"A{row}:{chr(64 + len(headers))}{row}")

    # Base case
    row = 2
    sheet.cell(row, 1).value = "Base"
    sheet.cell(row, 2).value = "50% fade"
    sheet.cell(row, 3).value = (
        assumptions.cogs_ex_da_pct_rev[0] + assumptions.sga_pct_rev[0]
        if assumptions.cogs_ex_da_pct_rev
        else 0.85
    )
    apply_percent(sheet, f"C{row}")
    sheet.cell(row, 4).value = assumptions.wacc
    apply_percent(sheet, f"D{row}")
    sheet.cell(row, 5).value = assumptions.terminal_growth_rate
    apply_percent(sheet, f"E{row}")

    # Bull case
    row = 3
    sheet.cell(row, 1).value = "Bull"
    sheet.cell(row, 2).value = "30% fade"
    sheet.cell(row, 3).value = (
        (assumptions.cogs_ex_da_pct_rev[0] + assumptions.sga_pct_rev[0] - 0.05)
        if assumptions.cogs_ex_da_pct_rev
        else 0.80
    )
    apply_percent(sheet, f"C{row}")
    sheet.cell(row, 4).value = assumptions.wacc - 0.01
    apply_percent(sheet, f"D{row}")
    sheet.cell(row, 5).value = assumptions.terminal_growth_rate + 0.005
    apply_percent(sheet, f"E{row}")

    # Bear case
    row = 4
    sheet.cell(row, 1).value = "Bear"
    sheet.cell(row, 2).value = "70% fade"
    sheet.cell(row, 3).value = (
        (assumptions.cogs_ex_da_pct_rev[0] + assumptions.sga_pct_rev[0] + 0.05)
        if assumptions.cogs_ex_da_pct_rev
        else 0.90
    )
    apply_percent(sheet, f"C{row}")
    sheet.cell(row, 4).value = assumptions.wacc + 0.01
    apply_percent(sheet, f"D{row}")
    sheet.cell(row, 5).value = assumptions.terminal_growth_rate - 0.005
    apply_percent(sheet, f"E{row}")

    set_column_widths(sheet, {1: 12, 2: 15, 3: 15, 4: 12, 5: 18})


def _write_wacc(sheet, assumptions) -> None:
    """Write WACC calculation to sheet."""
    row = 1
    sheet.cell(row, 1).value = "Component"
    sheet.cell(row, 2).value = "Value"
    apply_header(sheet, f"A{row}:B{row}")

    # Cost of Equity
    row = 2
    sheet.cell(row, 1).value = "Risk-Free Rate"
    sheet.cell(row, 2).value = assumptions.risk_free_rate
    apply_percent(sheet, f"B{row}")
    row += 1

    sheet.cell(row, 1).value = "Equity Risk Premium"
    sheet.cell(row, 2).value = assumptions.equity_risk_premium
    apply_percent(sheet, f"B{row}")
    row += 1

    sheet.cell(row, 1).value = "Beta"
    sheet.cell(row, 2).value = assumptions.beta
    row += 1

    cost_of_equity = assumptions.risk_free_rate + (assumptions.beta * assumptions.equity_risk_premium)
    sheet.cell(row, 1).value = "Cost of Equity"
    sheet.cell(row, 1).font = Font(bold=True)
    sheet.cell(row, 2).value = cost_of_equity
    sheet.cell(row, 2).font = Font(bold=True)
    apply_percent(sheet, f"B{row}")
    row += 2

    # Cost of Debt
    sheet.cell(row, 1).value = "Cost of Debt"
    sheet.cell(row, 2).value = assumptions.cost_of_debt
    apply_percent(sheet, f"B{row}")
    row += 1

    sheet.cell(row, 1).value = "Tax Rate"
    sheet.cell(row, 2).value = assumptions.tax_rate
    apply_percent(sheet, f"B{row}")
    row += 1

    after_tax_cost_of_debt = assumptions.cost_of_debt * (1 - assumptions.tax_rate)
    sheet.cell(row, 1).value = "After-Tax Cost of Debt"
    sheet.cell(row, 2).value = after_tax_cost_of_debt
    apply_percent(sheet, f"B{row}")
    row += 2

    # Weights
    sheet.cell(row, 1).value = "Debt/Equity Ratio"
    sheet.cell(row, 2).value = assumptions.debt_to_equity_ratio
    row += 1

    # Calculate weights
    debt_weight = assumptions.debt_to_equity_ratio / (1 + assumptions.debt_to_equity_ratio)
    equity_weight = 1 / (1 + assumptions.debt_to_equity_ratio)

    sheet.cell(row, 1).value = "Debt Weight"
    sheet.cell(row, 2).value = debt_weight
    apply_percent(sheet, f"B{row}")
    row += 1

    sheet.cell(row, 1).value = "Equity Weight"
    sheet.cell(row, 2).value = equity_weight
    apply_percent(sheet, f"B{row}")
    row += 2

    # WACC
    wacc = (equity_weight * cost_of_equity) + (debt_weight * after_tax_cost_of_debt)
    sheet.cell(row, 1).value = "WACC"
    sheet.cell(row, 1).font = Font(bold=True, size=12)
    sheet.cell(row, 2).value = wacc
    sheet.cell(row, 2).font = Font(bold=True, size=12)
    apply_percent(sheet, f"B{row}")

    set_column_widths(sheet, {1: 25, 2: 15})


def _write_sensitivities(sheet, sensitivity: dict) -> None:
    """Write 2D sensitivity table to sheet."""
    if not sensitivity:
        sheet.cell(1, 1).value = "No sensitivity data available"
        return

    # Get WACC and growth ranges
    wacc_values = sorted([float(k) for k in sensitivity.keys()])
    growth_values = sorted([float(k) for k in list(sensitivity.values())[0].keys()])

    # Headers
    row = 1
    sheet.cell(row, 1).value = "WACC \\ Terminal Growth"
    for col, growth in enumerate(growth_values, start=2):
        cell = sheet.cell(row, col)
        cell.value = f"{growth:.3f}"
    apply_header(sheet, f"A{row}:{chr(64 + len(growth_values) + 1)}{row}")

    # Data rows
    for row_idx, wacc in enumerate(wacc_values, start=2):
        cell = sheet.cell(row_idx, 1)
        cell.value = f"{wacc:.3f}"
        apply_header(sheet, f"A{row_idx}")

        wacc_key = f"{wacc:.3f}"
        if wacc_key in sensitivity:
            for col_idx, growth in enumerate(growth_values, start=2):
                growth_key = f"{growth:.3f}"
                if growth_key in sensitivity[wacc_key]:
                    sheet.cell(row_idx, col_idx).value = sensitivity[wacc_key][growth_key]
                    sheet.cell(row_idx, col_idx).number_format = '#,##0.00'

    set_column_widths(sheet, {1: 20, **{i: 12 for i in range(2, len(growth_values) + 2)}})


def _write_summary(sheet, valuation_output: ValuationOutput, ticker: str) -> None:
    """Write summary sheet with key outputs and price comparison."""
    row = _add_title_block(sheet, ticker, "Valuation Summary")

    results = valuation_output.results
    assumptions = valuation_output.assumptions

    # Valuation Outputs Box
    box_start_row = row
    sheet.cell(row, 1).value = "Valuation Outputs"
    sheet.cell(row, 1).font = Font(bold=True, size=12)
    row += 1

    summary_data = [
        ("Enterprise Value ($M)", to_millions(results.total_enterprise_value), "currency"),
        ("Net Debt ($M)", to_millions(assumptions.net_debt), "currency"),
        ("Equity Value ($M)", to_millions(results.equity_value), "currency"),
        ("Shares Outstanding (M)", to_millions(assumptions.shares_out), None),
        ("", "", None),  # Spacer
        ("Fair Value per Share", results.fair_value_per_share, "per_share"),
    ]

    for label, value, fmt_type in summary_data:
        if label:  # Skip spacer
            sheet.cell(row, 1).value = label
            sheet.cell(row, 2).value = value
            if fmt_type == "currency":
                apply_currency_millions(sheet, f"B{row}")
            elif fmt_type == "per_share":
                sheet.cell(row, 2).number_format = '#,##0.00'
                sheet.cell(row, 2).font = Font(bold=True, size=12)
        row += 1

    # Price Comparison Section
    row += 1
    sheet.cell(row, 1).value = "Price Comparison"
    sheet.cell(row, 1).font = Font(bold=True, size=12)
    row += 1

    # Fetch current price
    current_price = fetch_current_price(ticker)
    sheet.cell(row, 1).value = "Current Price"
    if current_price:
        sheet.cell(row, 2).value = current_price
    else:
        sheet.cell(row, 2).value = "N/A"
    sheet.cell(row, 2).number_format = '#,##0.00'
    row += 1

    sheet.cell(row, 1).value = "Fair Value per Share"
    sheet.cell(row, 2).value = results.fair_value_per_share
    sheet.cell(row, 2).number_format = '#,##0.00'
    sheet.cell(row, 2).font = Font(bold=True)
    row += 1

    sheet.cell(row, 1).value = "Implied Upside/(Downside)"
    if current_price and current_price > 0:
        upside = (results.fair_value_per_share / current_price) - 1.0
        sheet.cell(row, 2).value = upside
        if upside > 0:
            sheet.cell(row, 2).font = Font(bold=True, color="006100")  # Green for upside
        else:
            sheet.cell(row, 2).font = Font(bold=True, color="C00000")  # Red for downside
    else:
        sheet.cell(row, 2).value = "N/A"
    sheet.cell(row, 2).number_format = '0.0%'
    row += 1

    # Key Assumptions
    row += 1
    sheet.cell(row, 1).value = "Key Assumptions"
    sheet.cell(row, 1).font = Font(bold=True, size=12)
    row += 1

    assumptions_data = [
        ("WACC", assumptions.wacc, "percent"),
        ("Terminal Growth Rate", assumptions.terminal_growth_rate, "percent"),
        ("Tax Rate", assumptions.tax_rate, "percent"),
        ("Base Revenue ($M)", to_millions(assumptions.base_year_revenue), "currency"),
    ]

    for label, value, fmt_type in assumptions_data:
        sheet.cell(row, 1).value = label
        sheet.cell(row, 2).value = value
        if fmt_type == "currency":
            apply_currency_millions(sheet, f"B{row}")
        elif fmt_type == "percent":
            apply_percent(sheet, f"B{row}")
        row += 1

    # Apply input styling to assumptions
    apply_input(sheet, f"A{box_start_row + 1}:B{row - 1}")

    # Set column widths
    set_column_widths(sheet, {1: 30, 2: 18})
