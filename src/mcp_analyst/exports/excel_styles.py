"""Excel styling utilities for professional formatting."""

from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter


def header_style():
    """Get header cell style."""
    return {
        "font": Font(bold=True, color="FFFFFF", size=11),
        "fill": PatternFill(start_color="366092", end_color="366092", fill_type="solid"),
        "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        "border": Border(
            left=Side(style="thin", color="000000"),
            right=Side(style="thin", color="000000"),
            top=Side(style="thin", color="000000"),
            bottom=Side(style="medium", color="000000"),
        ),
    }


def input_style():
    """Get input cell style."""
    return {
        "fill": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
        "border": Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        ),
    }


def currency_millions_format():
    """Get currency format for $ Millions."""
    return '$#,##0'  # $43,978


def percent_format():
    """Get percent format."""
    return '0.0%'  # 15.2%


def per_share_format():
    """Get per-share format."""
    return '#,##0.00'  # 45.23


def apply_header(sheet, cell_range):
    """Apply header style to cell range."""
    from openpyxl.utils import range_boundaries
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    style = header_style()
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            cell = sheet.cell(row, col)
            for key, value in style.items():
                setattr(cell, key, value)


def apply_input(sheet, cell_range):
    """Apply input style to cell range."""
    from openpyxl.utils import range_boundaries
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    style = input_style()
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            cell = sheet.cell(row, col)
            for key, value in style.items():
                setattr(cell, key, value)


def apply_currency_millions(sheet, cell_range):
    """Apply $ Millions format to cell range."""
    from openpyxl.utils import range_boundaries
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    fmt = currency_millions_format()
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            cell = sheet.cell(row, col)
            cell.number_format = fmt


def apply_percent(sheet, cell_range):
    """Apply percent format to cell range."""
    from openpyxl.utils import range_boundaries
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    fmt = percent_format()
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            cell = sheet.cell(row, col)
            cell.number_format = fmt


def apply_number(sheet, cell_range, fmt='#,##0'):
    """Apply number format to cell range (for shares, counts, etc.)."""
    from openpyxl.utils import range_boundaries
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            cell = sheet.cell(row, col)
            cell.number_format = fmt


def apply_decimal(sheet, cell_range, fmt='0.000'):
    """Apply decimal format to cell range (for discount factors, etc.)."""
    from openpyxl.utils import range_boundaries
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            cell = sheet.cell(row, col)
            cell.number_format = fmt


def section_header_style():
    """Get section header style (dark fill, white font, merged)."""
    return {
        "font": Font(bold=True, color="FFFFFF", size=11),
        "fill": PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid"),
        "alignment": Alignment(horizontal="left", vertical="center"),
        "border": Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="medium"),
            bottom=Side(style="thin"),
        ),
    }


def apply_section_header(sheet, cell_range):
    """Apply section header style to cell range."""
    from openpyxl.utils import range_boundaries
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    style = section_header_style()
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            cell = sheet.cell(row, col)
            for key, value in style.items():
                setattr(cell, key, value)


def set_column_widths(sheet, widths: dict):
    """Set column widths.
    
    Args:
        widths: Dict mapping column letters or indices to widths
    """
    for col, width in widths.items():
        if isinstance(col, int):
            col_letter = get_column_letter(col)
        else:
            col_letter = col
        sheet.column_dimensions[col_letter].width = width


def to_millions(value: float) -> float:
    """Convert value to millions."""
    return value / 1_000_000


def freeze_panes(sheet, cell: str):
    """Freeze panes at specified cell."""
    sheet.freeze_panes = cell

